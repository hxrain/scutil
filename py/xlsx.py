import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook

from hash_calc import *

'''
excel文件生成器,每次生成都会覆盖原有内容

xw = xlsx_writer('tst.xlsx')
xw.create('表格1', ['列1', '列2'])
xw.append(('v10', 'v11'))
xw.appendx([('v20', 'v21'),('v30', 'v31')])

xw.create('表格2', ['列1', '列2'])
xw.append(('v10', 'v11'))
xw.append(('v11', 'v12'))
xw.close()

'''


class xlsx_sheet:
    """excel数据页功能封装,用于创建新内容数据"""

    def __init__(self, keyIdx=0):
        self.sheet = None
        self.rows = 0
        self.keys = None
        self.keyIdx = keyIdx

    def append(self, t, fmts=None):
        """追加一个元组"""
        if not self.sheet:
            return -1

        if self.keys:
            key = calc_key(t, self.keyIdx)
            if key in self.keys:
                return 0
            self.keys.add(key)

        if fmts is None:
            fmts = [0] * len(t)

        self.rows += 1
        for col in range(len(t)):
            self.sheet.write(self.rows, col, t[col], self.cell_style[fmts[col]])

        return 2

    def appendx(self, lst):
        """追加元组列表"""
        for t in lst:
            self.append(t)


class xlsx_maker:
    """excel数据页生成器功能封装,用于创建多个数据页"""

    def __init__(self, fname):
        # 打开文件
        self.book = xlsxwriter.Workbook(fname)
        # 定义单元格样式
        self.cell_style = []
        self.add_style({
            'text_wrap': False,
            'valign': 'vcenter',
            'align': 'left',
        })

    def add_style(self, styles_dict):
        fmt = self.book.add_format(styles_dict)
        self.cell_style.append(fmt)
        return len(self.cell_style) - 1

    def create(self, sheet_name, cols=None, chk_keys=True, keyIdx=None):
        """创建数据表,告知表名与列头"""
        s = xlsx_sheet(keyIdx)
        s.sheet = self.book.add_worksheet(sheet_name)
        s.cell_style = self.cell_style

        if chk_keys:
            s.keys = set()

        # 添加列头
        if cols:
            s.rows = 0
            for i in range(len(cols)):
                s.sheet.write(s.rows, i, cols[i])
        else:
            s.rows = -1
        return s

    def close(self):
        """关闭文档保存内容"""
        self.book.close()


class xlsx_writer:
    """excel单页书写器功能封装,结合数据页生成器与数据页功能封装,简化使用."""

    def __init__(self, fname, keyIdx=0):
        self.maker = xlsx_maker(fname)
        self.sheet = None

    def add_style(self, styles_dict):
        self.maker.add_style(styles_dict)

    def create(self, sheet_name, cols=None, chk_keys=True):
        """创建数据表,告知表名与列头"""
        self.sheet = self.maker.create(sheet_name, cols, chk_keys)

    def append(self, t, fmts=None):
        """追加一个元组"""
        return self.sheet.append(t, fmts)

    def appendx(self, lst):
        """追加元组列表"""
        return self.sheet.appendx(t)

    def close(self):
        """关闭文档保存内容"""
        if not self.sheet:
            return
        self.maker.close()
        self.sheet = None


class xlsx_editor:
    """excel文件读取编辑器,未做错误处理,使用时需做异常捕获"""

    def __init__(self, fname=None, data_only=True):
        """构造并装载数据,告知是否取公式的数据结果"""
        self.file = None
        if fname != '':
            self.open(fname, data_only)

    def open(self, fname=None, data_only=True):
        """打开文件装载数据/或创建新文档,告知是否取公式的数据结果"""
        try:
            if fname:
                self.file = load_workbook(fname, data_only=data_only)
                self.fname = fname
            else:
                self.file = Workbook()
                self.fname = None
            return ''
        except Exception as e:
            return str(e)

    def is_opened(self):
        return self.file is not None

    def sheets(self):
        """获取当前excel中每个表格页的名称.返回值:[tab页名称列表]"""
        return self.file.sheetnames

    def cols(self, sheet_idx=0):
        """获取指定tab页含有的数据列数"""
        sheet = self.get_sheet(sheet_idx)  # 按索引得到指定tab页
        return sheet.max_column

    def rows(self, sheet_idx=0):
        """获取指定tab页含有的数据行数"""
        sheet = self.get_sheet(sheet_idx)  # 按索引得到指定tab页
        return sheet.max_row

    def get_sheet(self, sheet_idx):
        """获取指定tab页对象:"""
        if isinstance(sheet_idx, int):
            return self.file.worksheets[sheet_idx]  # 按索引得到指定tab页
        else:
            return self.file[sheet_idx]  # 按名称得到指定tab页

    def new_sheet(self, title, new_idx=None):
        """创建新tab页,告知标题,指定新tab页的索引位置(默认为最后)"""
        self.file.create_chartsheet(title, new_idx)

    def set_line(self, row, vals, sheet_idx=0):
        """给指定tab页的指定行单元格写数据.row/col行列计数从0开始."""
        sheet = self.get_sheet(sheet_idx)  # 按索引得到指定tab页
        for col, val in enumerate(vals):
            sheet.cell(row=row + 1, column=col + 1).value = val

    def get_line(self, row, sheet_idx=0, cols=0):
        """从指定tab页的指定行row获取指定列数量的数据.row行计数从0开始."""
        if cols == 0:
            cols = self.cols(sheet_idx)
        sheet = self.get_sheet(sheet_idx)  # 按索引得到指定tab页
        return [sheet.cell(row=row + 1, column=col + 1).value for col in range(cols)]

    def set_cell(self, row, col, val, sheet_idx=0):
        """给指定tab页的指定行列单元格写数据.row/col行列计数从0开始."""
        sheet = self.get_sheet(sheet_idx)  # 按索引得到指定tab页
        sheet.cell(row=row + 1, column=col + 1).value = val

    def get_cell(self, row, col, sheet_idx=0):
        """获取指定行列单元格的值,row/col行列计数从0开始."""
        sheet = self.get_sheet(sheet_idx)  # 按索引得到指定tab页
        return sheet.cell(row=row + 1, column=col + 1).value

    def append(self, line, sheet_idx=0):
        """给指定tab页追加一行数据(只能从第二行开始追加.首行保留)"""
        sheet = self.get_sheet(sheet_idx)  # 按索引得到指定tab页
        sheet.append(line)

    def query(self, row=None, col=None, sheet_idx=0, looper=None):
        """获取指定tab页中指定行列范围的数据.row/col行列计数从0开始.返回值:[(),(),...]列表,行列数据;[(None,)]代表tab页为空."""

        class loop:
            """定义内置的遍历处理器,用于累积输出结果."""

            def __init__(self):
                self.rst = []

            def __call__(self, *args, **kwargs):
                self.rst.append(args[0])

            def result(self):
                return self.rst

        if looper is None:
            looper = loop()  # 使用内置的遍历处理器

        sheet = self.get_sheet(sheet_idx)  # 按索引得到指定tab页

        # 调教row参数变为行范围,row为列表[]则代表指定的具体行
        if row is None:
            row = range(sheet.max_row)  # 全部行的范围
        elif isinstance(row, int):
            row = [row]  # 单行
        elif isinstance(row, tuple):
            row = range(row[0], row[1] if row[1] != -1 else self.rows(sheet_idx))  # 元组代表开始和结束行的范围

        # 调教col参数变为列范围,col为列表[]则代表指定的具体列
        if col is None:
            col = range(sheet.max_column)  # 全部列的范围
        elif isinstance(col, int):
            col = [col]  # 单列
        elif isinstance(col, tuple):
            col = range(col[0], col[1] if col[1] != -1 else self.cols(sheet_idx))  # 元组代表开始和结束列的范围

        # 进行行列读取,得到每个格子的数据
        rows = list(sheet.rows)
        for rowi in row:
            cols = rows[rowi]
            vals = tuple(cols[coli].value for coli in col)
            looper(vals)

        return looper.result()

    def save(self, fname=None):
        """保存修改或另存为新文件"""
        if fname is None:
            fname = self.fname
        self.file.save(fname)

    def close(self):
        """关闭编辑器"""
        self.file.close()
        self.file = None
